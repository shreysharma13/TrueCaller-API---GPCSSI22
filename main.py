########################################################################################
######################          Import packages      ###################################
########################################################################################
from flask import Blueprint, send_file, url_for, render_template, flash
from flask_login import login_required, current_user
from __init__ import create_app, db
from werkzeug.utils import secure_filename
import pandas as pd
from xml.dom import minidom
import openpyxl as xl
import cgi, os
import cgitb; cgitb.enable()
from flask import Flask, render_template, redirect, url_for, request
from openpyxl.utils.exceptions import InvalidFileException
########################################################################################
#our main blueprint
main = Blueprint('main', __name__)

@main.route('/') # home page that return 'index'
def index():
    return render_template('index.html')

@main.route('/profile') # profile page that return 'profile'
@login_required
def profile():
    return render_template('profile.html', name=current_user.name)

@main.route('/data', methods = ['GET', 'POST'])

def data():

    if request.method == 'POST' and request.form['submit'] == 'view':
        f = request.files['upload']

        try:
            f.save(secure_filename(f.filename))
            print(f.filename)
            wb = xl.load_workbook(f.filename)
            sheet = wb['Sheet1']
            data= pd.read_excel(f.filename)
            return render_template('view.html' , data =data.to_html())

        except InvalidFileException:
           data= "OOPS!Wrong file extension..please upload .xlsx file"
           return render_template('view.html' , data =data)

        except FileNotFoundError:
            data = "oops! Cant find file"
            return render_template('view.html' , data =data)




    if request.method == 'POST' and request.form['submit'] == 'submit':
        f = request.files['upload']

        try:
            f.save(secure_filename(f.filename))
            fname = f.filename
            print(f.filename)
            wb = xl.load_workbook(f.filename)
            sheet = wb['Sheet1']

        except FileNotFoundError:
            data = "oops! Cant find file"
            return render_template('view.html' , data =data)
        except InvalidFileException:
            data = "OOPS!Wrong file extension..please upload .xlsx file"
            return render_template('view.html' , data =data)

        def cellEntry(row, column, attribute ):
            new_cell = sheet.cell(row, column)
            new_cell.value = str(attribute)
            #enters cell data in specific cell, attribute is the data entered

        def trial():
            #takes input str and the row where the corresponding data has to be printed

            for row in range(2, sheet.max_row+1):
                data = sheet.cell(int(row), 1)
                print(str(type(data.value)))
                if(str(type(data.value)) != "<class 'int'>"):
                    cellEntry(row, 2 , "invalid character")
                else:
                    if(len(str(data.value)) != 10):
                        cellEntry(row, 2 , "invalid number")
                    else:

                        # data = whatever is in the sheet at row,column specified...in this case it would be our phoennumber
                        cmd = "truecallerjs -s " + str(data.value) + " --xml | sed '1d'"
                        input = os.popen(cmd).read()
                        p3: None = minidom.parseString(input)
                        print(type(p3))
                        #chek type of p3 LINE NOT NEEDED
                        def getTagElement(tag_name, row , column):
                            #gets tag name and prints the tag data in the cell specified
                            try:

                                if (tag_name=='id'):
                                    element = p3.getElementsByTagName(tag_name)
                                    print(element[1].firstChild.data)
                                    cellEntry(row,  column, element[1].firstChild.data)
                                else:
                                    #if tag is not id print element [0] ie. the first instance of the tag
                                    element = p3.getElementsByTagName(tag_name)
                                    #gets of data all instances of the tag in an array element
                                    cellEntry(row,  column, element[0].firstChild.data)
                                    #enters it in the excel sheet


                            except IndexError:
                                cellEntry(row,  column, "null")
                        try:
                            if (p3.getElementsByTagName("errorResp")[0].firstChild.data =="Request failed with status code 429"):
                                cellEntry(row,2, "Rate Limit has been reached")
                                term="open -a Terminal"
                                #POPS A TERMINAL WINDOW FOR TRUECALLER LOGIN PHONE NUMBER VERIFICATION.
                                #TERMINAL COMMANDS TO BE USED - truecallerjs login
                                #### enter phone number and get otp verification.
                                os.popen(term)
                                ##### LINE 109-111 ARE FOR UNIX BASED SYSTEMS, FOR WINDOWS AND LINUX SYSTEMS,
                                #####  CORRESPONDING ALTERNATES ARE AVAILABLE.
                            elif (p3.getElementsByTagName("errorResp")[0].firstChild.data =="Request failed with status code 401"):
                                cellEntry(row,2, "Lost authentication")
                                term="open -a Terminal"
                                #POPS A TERMINAL WINDOW FOR TRUECALLER LOGIN PHONE NUMBER VERIFICATION.
                                #TERMINAL COMMANDS TO BE USED - truecallerjs login
                                #### enter phone number and get otp verification.
                                os.popen(term)
                                ##### LINE 109-111 ARE FOR UNIX BASED SYSTEMS, FOR WINDOWS AND LINUX SYSTEMS,
                                #####  CORRESPONDING ALTERNATES ARE AVAILABLE.
                        except IndexError :

                            getTagElement("name" , row , 2)
                            cellEntry(1,2,"Name")
                            # cellEntry(1,3,"e164Format")
                            # getTagElement("e164Format" ,row, 3)
                            getTagElement("carrier" ,row, 3)
                            cellEntry(1,3,"Carrier")
                            getTagElement("city" , row, 4)
                            cellEntry(1,4,"City")
                            getTagElement("image" , row , 5)
                            cellEntry(1,5,"Image")
                            getTagElement("id" , row , 6)
                            cellEntry(1,6,"ID")
                            getTagElement("caption" , row, 7)
                            cellEntry(1,7,"Caption")
                            # getTagElement("altName", row , 8)
                            # cellEntry(1,8,"AltName")
                            # getTagElement("imID", row , 9)
                            # cellEntry(1,9,"imID")
                            ##additional data which is optional

        trial()
        wb.save('result.xlsx')
        data= pd.read_excel('result.xlsx')
        return render_template('data.html' , data =data.to_html())

@main.route('/download')
def download():
         path = 'result.xlsx'
         return send_file(path, as_attachment=True)


app = create_app() # we initialize our flask app using the __init__.py function
if __name__ == '__main__':
    db.create_all(app=create_app()) # create the SQLite database
    app.run(debug=True) # run the flask app on debug mode
